"""
Build the URL Decisions spreadsheet for the dgtlgroup.io WordPress migration.

Reads from the audit at docs/audit/wordpress-audit.md and produces an .xlsx
where Will can review every existing URL and check the ones to keep.

Default: every row is UNCHECKED. Will checks the keepers; everything still
unchecked at the end gets a 301 (or 410) in the migration redirect map.

Run from repo root:
    python docs/audit/build-url-decisions.py
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

OUTPUT = "docs/audit/url-decisions.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Data — every URL discovered in the audit, with recommendation
#
# Recommendation codes:
#   KEEP    — clearly migrate this page to the new site
#   DROP    — page should not exist on the new site (301 or 410)
#   DECIDE  — needs Will's call before we can map it
#   MERGE   — fold into another page
# ─────────────────────────────────────────────────────────────────────────────

ROWS = [
    # ── Bucket A — Core marketing pages ─────────────────────────────────────
    ("Core marketing", "/",
     "Homepage",
     "Hero, services overview, featured projects, client logos, testimonials, newsletter signup",
     "KEEP", "/"),
    ("Core marketing", "/about-us/",
     "About Us",
     "Company story, '75+ projects' stat (other stat fields show literal 0% / $0+ placeholders), client logo wall",
     "KEEP", "/about"),
    ("Core marketing", "/services/",
     "Services",
     "Lists 6 services as cards: Content Creation, SMM, Influencer, PR, Web Design & Dev, Graphics. No individual service detail pages exist",
     "KEEP", "/services"),
    ("Core marketing", "/work/",
     "Work / Portfolio",
     "Portfolio grid linking to /work/<slug>/ case studies (and one orphan root URL)",
     "KEEP", "/work"),
    ("Core marketing", "/contact-us/",
     "Contact Us",
     "Contact form, contact@dgtlgroup.io, +1 (213) 772-6886 / +1 (647) 930-4443, 487 Adelaide St. W Toronto",
     "KEEP", "/contact"),

    # ── Bucket B — Secondary marketing / utility pages ──────────────────────
    ("Secondary", "/blog/",
     "Blog",
     "Index page — links to root-URL articles (NOT to /insights/)",
     "KEEP", "/blog"),
    ("Secondary", "/insights/",
     "Insights archive",
     "Custom post type archive — links to /insights/<slug>/ stub pages. Direct duplicate of /blog/",
     "DROP", "301 → /blog"),
    ("Secondary", "/join-dgtl-team/",
     "Join DGTL Team (Careers)",
     "Careers page; lists 2 SMM jobs in Indonesia + Philippines but body text also says 'no opening'",
     "KEEP", "/careers"),
    ("Secondary", "/join-dgtl-influence/",
     "Join DGTL Influence",
     "Influencer recruiting landing page. H1 ends mid-sentence with a literal | character (unfinished draft)",
     "DECIDE", "/influence  or merge into /careers"),

    # ── Bucket C — Case studies under /work/ ────────────────────────────────
    ("Case study", "/work/art-villas-costa-rica-content-campaign/",
     "Art Villas Costa Rica Content Campaign",
     "Costa Rica luxury villas — Photography, Videography, UGC",
     "KEEP", "/work/art-villas-costa-rica"),
    ("Case study", "/work/six-senses-content-campaign/",
     "Six Senses Ibiza Content Campaign",
     "Six Senses Ibiza resort — Photography, Videography, FPV Drone, Influencer",
     "KEEP", "/work/six-senses-ibiza"),
    ("Case study", "/work/pacific-high-dewata-content-campaign/",
     "Pacific High Dewata Content Campaign",
     "Pacific High Indonesia — Photography, Videography, Drone, Underwater",
     "KEEP", "/work/pacific-high-dewata"),
    ("Case study", "/work/nebula-projector/",
     "Anker Nebula Projector",
     "Anker — Photography, Videography, UGC. Slug too generic; new slug should include client",
     "KEEP", "/work/anker-nebula-projector"),
    ("Case study", "/work/canon/",
     "Canon R10 Launch Campaign",
     "Canon Canada — Photography, content campaign. Slug is just the brand name today",
     "KEEP", "/work/canon-r10-launch"),
    ("Case study", "/work/sounds-of-the-city-epidemic-sound-campaign/",
     "Sounds of the City × Epidemic Sound",
     "Epidemic Sound campaign — Photography, Videography, VFX",
     "KEEP", "/work/epidemic-sound-sounds-of-the-city"),
    ("Case study", "/work/epidemicsound/",
     "Epidemic Sound Social Media Management",
     "Epidemic Sound — UGC, Videography, Social Media Marketing. 12+ million views (only case with a numeric result)",
     "KEEP", "/work/epidemic-sound-social"),
    ("Case study", "/work/a-day-with-swae-lee/",
     "A Day with Swae Lee at Cabana & Rebel",
     "Swae Lee Toronto event — Photography, Videography, Drone, FPV",
     "KEEP", "/work/swae-lee-toronto"),
    ("Case study", "/work/jassa-dhillon-you-music-video/",
     "Jassa Dhillon 'YOU' Music Video",
     "Music video production — Photography, Videography",
     "KEEP", "/work/jassa-dhillon-you"),
    ("Case study", "/work/on-running-train-on-clouds/",
     "ON Running — Train on Clouds",
     "ON Running commercial — Videography, Photography",
     "KEEP", "/work/on-running-train-on-clouds"),
    ("Case study", "/work/spider-man-raf-film/",
     "Spider-Man: RAF",
     "Carlos Juico fan film — Feature Film, Videography, Influencer Marketing",
     "KEEP", "/work/spider-man-raf"),
    ("Case study", "/work/flockamg/",
     "Flocka MG",
     "Flocka MG (artist) — Music, Fashion, Music Video, DMTV",
     "KEEP", "/work/flocka-mg"),
    ("Case study", "/work/lum/",
     "LUM (case study)",
     "LUM artist case study — also exists as /lum/ talent profile (slug collision — pick one canonical URL)",
     "DECIDE", "/work/lum  (collides with /lum/)"),

    # ── Bucket D — Talent / influencer / artist profile pages (orphan root URLs)
    ("Talent / sub-brand", "/dom/",
     "Dom Vallie",
     "Canadian rapper, JUNO nominee, 75k+ Spotify monthly listeners. External talent represented by DGTL Influence",
     "DECIDE", "/influence/dom-vallie"),
    ("Talent / sub-brand", "/lum/",
     "LUM (talent profile)",
     "Artist — also has /work/lum/ case study (slug collision)",
     "DECIDE", "/influence/lum  or fold into /work/lum"),
    ("Talent / sub-brand", "/tye/",
     "Tye",
     "Toronto fashion photographer & creative director",
     "DECIDE", "/influence/tye"),
    ("Talent / sub-brand", "/stacy/",
     "Anastasiia",
     "Photographer, social media manager, model — explicitly described as working FOR DGTL Group (likely staff, not external talent)",
     "DECIDE", "/team/anastasiia"),
    ("Talent / sub-brand", "/theburnstwins/",
     "The Burns Twins",
     "Toronto identical-twin prank/comedy creators",
     "DECIDE", "/influence/the-burns-twins"),
    ("Talent / sub-brand", "/kamil/",
     "Kamil Galin",
     "Director of Photography at DGTL (likely staff, not external talent). Hosts 'B0:0M' short film",
     "DECIDE", "/team/kamil-galin"),
    ("Talent / sub-brand", "/5a1ive/",
     "5a1ive",
     "Toronto streetwear designer/creator",
     "DECIDE", "/influence/5a1ive"),
    ("Talent / sub-brand", "/shane/",
     "Shane Boyer",
     "American skateboarder/influencer",
     "DECIDE", "/influence/shane-boyer"),
    ("Talent / sub-brand", "/dmtv/",
     "DMTV",
     "Music media collective sub-brand. Aggregates 5+ music video projects (Nick Souza, FBK, CAIRO!, Lil Esso, Papi AQ) under one URL instead of giving each its own slug",
     "DECIDE", "/dmtv  or break each video into /work/<slug>"),
    ("Talent / sub-brand", "/dgtltags/",
     "Digital NFC Tags",
     "Product page for retail NFC tags. Looks like a separate, unrelated business line",
     "DROP", "Move to separate domain or 410"),
    ("Talent / sub-brand", "/lil-tjay-calgary-live-recap/",
     "Lil Tjay — Calgary Live Recap",
     "Case-study-style content sitting at root URL. The /work/ index links to a non-existent /work/lil-tjay-calgary-live-recap/ for the same content",
     "KEEP", "/work/lil-tjay-calgary-live-recap"),

    # ── Bucket E — Real blog posts at root URLs ─────────────────────────────
    ("Blog post (real)", "/maximizing-your-digital-marketing-strategy/",
     "Maximizing Your Digital Marketing Strategy",
     "Full article, 6 sections. AI-generated DALL·E hero image",
     "KEEP", "/blog/maximizing-your-digital-marketing-strategy"),
    ("Blog post (real)", "/understanding-consumer-behavior-in-2025/",
     "Understanding Consumer Behavior in 2025",
     "Full article, 7 sections. AI-generated DALL·E hero image",
     "KEEP", "/blog/understanding-consumer-behavior-in-2025"),
    ("Blog post (real)", "/essentials-when-building-new-website-ui-ux/",
     "5 UI/UX Essentials When Building a New Website",
     "Full article, 5 sections. AI-generated DALL·E hero image",
     "KEEP", "/blog/ui-ux-essentials-new-website"),

    # ── Bucket F — /insights/ legacy stub posts ─────────────────────────────
    ("Insights stub", "/insights/maximizing-your-digital-marketing-strategy/",
     "Maximizing Your Digital Marketing Strategy (stub)",
     "Stub page — chrome only. Same article as the root URL version above (duplicate)",
     "DROP", "301 → /blog/maximizing-your-digital-marketing-strategy"),
    ("Insights stub", "/insights/the-future-of-digital-marketing/",
     "The Future of Digital Marketing (stub)",
     "Stub only. NO real content anywhere on the site. Was the article ever written?",
     "DECIDE", "Kill (410) or write the article"),
    ("Insights stub", "/insights/understanding-consumer-behavior-in-2024/",
     "Understanding Consumer Behavior in 2024 (stub)",
     "Stub only. Probably superseded by the 2025 version",
     "DROP", "301 → /blog/understanding-consumer-behavior-in-2025"),

    # ── Bucket G — WordPress category archives ──────────────────────────────
    ("Category archive", "/category/marketing/",
     "Marketing category archive",
     "WordPress-generated category page, 1 post",
     "DROP", "301 → /blog"),
    ("Category archive", "/category/web/",
     "Web category archive",
     "WordPress-generated category page, 1 post",
     "DROP", "301 → /blog"),
    ("Category archive", "/category/content-creation/",
     "Content Creation category archive",
     "WordPress-generated category page, 1 post",
     "DROP", "301 → /blog"),
    ("Category archive", "/category/sales/",
     "Sales category archive",
     "WordPress-generated category page, 1 post",
     "DROP", "301 → /blog"),
    ("Category archive", "/category/digital-marketing/",
     "Digital Marketing category archive",
     "WordPress-generated category page, 1 post",
     "DROP", "301 → /blog"),
    ("Category archive", "/category/ai/",
     "AI category archive",
     "WordPress-generated category page, 1 post",
     "DROP", "301 → /blog"),

    # ── Bucket H — CPT / placeholder pages indexed in production ────────────
    ("Placeholder / CPT", "/services-lists/service-title-goes-here/",
     "Service Title Goes Here",
     "Lorem-ipsum placeholder. Indexed in services-list-sitemap.xml",
     "DROP", "410 Gone"),
    ("Placeholder / CPT", "/services-lists/service-title-goes-here-2/",
     "Service Title Goes Here (#2)",
     "Lorem-ipsum placeholder #2. Indexed in services-list-sitemap.xml",
     "DROP", "410 Gone"),
    ("Placeholder / CPT", "/client-feedback/",
     "Client Feedback archive",
     "Empty CPT archive",
     "DROP", "301 → /about"),
    ("Placeholder / CPT", "/client-feedback/filip-zak/",
     "Client Feedback — Filip Žák",
     "Empty CPT entry — chrome only. Real testimonial quote lives in homepage HTML, not here",
     "DROP", "301 → /about"),
    ("Placeholder / CPT", "/client-feedback/jane-doe/",
     "Client Feedback — Jane Doe",
     "Lorem-ipsum placeholder",
     "DROP", "410 Gone"),
    ("Placeholder / CPT", "/testimonial/",
     "Testimonial archive",
     "Empty CPT archive",
     "DROP", "301 → /about"),
    ("Placeholder / CPT", "/testimonial/filip_zak/",
     "Testimonial — Filip Žák",
     "Empty CPT entry — chrome only",
     "DROP", "301 → /about"),
    ("Placeholder / CPT", "/testimonial/name-surname-4/",
     "Testimonial — Name Surname 4",
     "Lorem-ipsum placeholder",
     "DROP", "410 Gone"),

    # ── Bucket I — Legal ────────────────────────────────────────────────────
    ("Legal", "/privacy-policy/",
     "Privacy Policy",
     "Privacy policy from 'DGTL Group Holdings Limited'. Refers users to 'primary email address specified in your account'",
     "KEEP", "/legal/privacy  (needs legal sign-off — entity / jurisdiction inconsistencies)"),
    ("Legal", "/terms-of-use/",
     "Terms of Use",
     "Site terms; governed by Saint-Kitts & Nevis law; legal contact legal@dgtlgroup.io",
     "KEEP", "/legal/terms  (needs legal sign-off)"),

    # ── Bucket J — Test / staging ───────────────────────────────────────────
    ("Test / staging", "/testing-page/",
     "Testing Page",
     "Test/placeholder page that should NOT be in production. Currently indexed in page-sitemap.xml",
     "DROP", "410 Gone"),

    # ── Bucket K — Files / external ─────────────────────────────────────────
    ("File / external", "/locations.kml",
     "locations.kml",
     "Geographic data file referenced by local-sitemap.xml. Local SEO setup attempted but not finished",
     "DROP", "410 (regenerate from Payload if needed)"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Build the workbook
# ─────────────────────────────────────────────────────────────────────────────

wb = Workbook()

# ── Sheet 1: URL Decisions ──────────────────────────────────────────────────
ws = wb.active
ws.title = "URL Decisions"

HEADERS = [
    "Keep?",
    "Category",
    "Old URL",
    "Title",
    "Description",
    "Recommendation",
    "Suggested new URL / action",
]

# Header row
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="0A0A0A")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

for col_idx, label in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

ws.row_dimensions[1].height = 28

# Body rows
body_font = Font(name="Arial", size=10)
body_align = Alignment(vertical="top", wrap_text=True)
checkbox_align = Alignment(horizontal="center", vertical="center")
mono_font = Font(name="Consolas", size=10)
thin = Side(border_style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, (category, url, title, desc, rec, new_url) in enumerate(ROWS, start=2):
    # A — checkbox cell (boolean, default FALSE)
    a = ws.cell(row=i, column=1, value=False)
    a.alignment = checkbox_align
    a.font = Font(name="Arial", size=14)
    # Custom format that displays ☑ for TRUE and ☐ for FALSE
    a.number_format = '[=TRUE]"\u2611";[=FALSE]"\u2610"'
    a.border = border

    # B — Category
    b = ws.cell(row=i, column=2, value=category)
    b.font = body_font
    b.alignment = Alignment(vertical="top", wrap_text=True)
    b.border = border

    # C — Old URL (monospace for readability)
    c = ws.cell(row=i, column=3, value=url)
    c.font = mono_font
    c.alignment = Alignment(vertical="top")
    c.border = border

    # D — Title
    d = ws.cell(row=i, column=4, value=title)
    d.font = Font(name="Arial", size=10, bold=True)
    d.alignment = body_align
    d.border = border

    # E — Description
    e = ws.cell(row=i, column=5, value=desc)
    e.font = body_font
    e.alignment = body_align
    e.border = border

    # F — Recommendation (color-coded fill set per row below via conditional fmt)
    f = ws.cell(row=i, column=6, value=rec)
    f.font = Font(name="Arial", size=10, bold=True)
    f.alignment = Alignment(horizontal="center", vertical="center")
    f.border = border

    # G — Suggested new URL or action
    g = ws.cell(row=i, column=7, value=new_url)
    g.font = mono_font
    g.alignment = Alignment(vertical="top", wrap_text=True)
    g.border = border

    # Row height tuned for wrapped descriptions
    ws.row_dimensions[i].height = 38

# Column widths
ws.column_dimensions["A"].width = 8     # Keep?
ws.column_dimensions["B"].width = 18    # Category
ws.column_dimensions["C"].width = 52    # Old URL
ws.column_dimensions["D"].width = 36    # Title
ws.column_dimensions["E"].width = 60    # Description
ws.column_dimensions["F"].width = 16    # Recommendation
ws.column_dimensions["G"].width = 48    # Suggested new URL

# Freeze the header row
ws.freeze_panes = "A2"

# Auto-filter on the whole table
last_row = len(ROWS) + 1
last_col_letter = get_column_letter(len(HEADERS))
ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

# Data validation: column A is TRUE/FALSE only
dv_keep = DataValidation(
    type="list",
    formula1='"TRUE,FALSE"',
    allow_blank=False,
    showDropDown=False,  # leave dropdown arrow visible (Excel inverts this flag)
    showErrorMessage=True,
    errorTitle="Pick TRUE or FALSE",
    error="Click the cell, then pick TRUE (☑ keep) or FALSE (☐ drop) from the dropdown.",
)
dv_keep.add(f"A2:A{last_row}")
ws.add_data_validation(dv_keep)

# Conditional formatting on the Recommendation column (F)
def add_rec_rule(value, fill_color, font_color="000000"):
    rule = FormulaRule(
        formula=[f'$F2="{value}"'],
        fill=PatternFill("solid", start_color=fill_color),
        font=Font(name="Arial", size=10, bold=True, color=font_color),
    )
    ws.conditional_formatting.add(f"F2:F{last_row}", rule)

add_rec_rule("KEEP",   "C6F6D5", "1F6B3F")  # green
add_rec_rule("DROP",   "FAD2CF", "8B1A10")  # red
add_rec_rule("DECIDE", "FFE9B0", "8C5A00")  # amber
add_rec_rule("MERGE",  "D9E3FF", "1F3A8A")  # blue

# Conditional formatting on the whole row when Keep? = TRUE
keep_rule = FormulaRule(
    formula=[f'$A2=TRUE'],
    fill=PatternFill("solid", start_color="EEFCEF"),
)
ws.conditional_formatting.add(f"A2:G{last_row}", keep_rule)

# Header comment on Keep? explaining the workflow
ws["A1"].comment = Comment(
    "Default: every row is unchecked (FALSE / ☐).\n\n"
    "Click a cell, pick TRUE from the dropdown, and the cell becomes ☑.\n\n"
    "When you're done, every row still showing ☐ becomes a 301 (or 410) "
    "in the migration redirect map.",
    "Claude",
)

# ── Sheet 2: Legend ─────────────────────────────────────────────────────────
legend = wb.create_sheet("Legend")
legend.column_dimensions["A"].width = 20
legend.column_dimensions["B"].width = 90

legend_rows = [
    ("How this sheet works", ""),
    ("",
     "Every URL on the current dgtlgroup.io WordPress site is one row on the URL Decisions tab. "
     "Default state for every row is UNCHECKED (☐). Tick the box for every page you want to "
     "migrate to the new Next.js + Payload site. Anything still unchecked at the end becomes a "
     "301 redirect (or a 410 Gone for placeholder/junk pages) in the migration redirect map."),
    ("", ""),
    ("How to check a row", ""),
    ("",
     "Click the cell in the Keep? column. A small dropdown arrow appears — pick TRUE. The cell "
     "flips from ☐ to ☑ and the whole row turns light green."),
    ("", ""),
    ("Recommendation values (column F)", ""),
    ("KEEP",   "Clearly migrate this page to the new site. Default action: tick the box."),
    ("DROP",   "Page should not exist on the new site. Default action: leave unchecked → 301 or 410."),
    ("DECIDE", "Needs your call — the audit couldn't make a confident recommendation."),
    ("MERGE",  "Should be folded into another page (rare in this set)."),
    ("", ""),
    ("Categories (column B)", ""),
    ("Core marketing",     "Top-level marketing pages (home, about, services, work, contact)."),
    ("Secondary",          "Secondary marketing pages (blog index, careers, etc.)."),
    ("Case study",         "Pages under /work/<slug>/."),
    ("Talent / sub-brand", "Orphan root URLs that profile staff, external talent, or sub-brands (DMTV, DGTL Tags)."),
    ("Blog post (real)",   "Real blog articles that live at root URLs (not under /blog/ or /insights/)."),
    ("Insights stub",      "Empty stub pages under /insights/. Most are duplicates of the real blog posts."),
    ("Category archive",   "WordPress-generated /category/<name>/ pages — one post each."),
    ("Placeholder / CPT",  "Lorem-ipsum or empty custom-post-type entries that leaked into production."),
    ("Legal",              "Privacy policy and terms of use pages."),
    ("Test / staging",     "Test/draft pages that were left publicly indexed."),
    ("File / external",    "Non-page files referenced from sitemaps."),
    ("", ""),
    ("Total URLs", str(len(ROWS))),
    ("Audit source", "docs/audit/wordpress-audit.md (commit 06da6ca)"),
    ("Generated by", "docs/audit/build-url-decisions.py"),
]

for i, (label, body) in enumerate(legend_rows, start=1):
    a = legend.cell(row=i, column=1, value=label)
    b = legend.cell(row=i, column=2, value=body)
    if not body:
        a.font = Font(name="Arial", size=12, bold=True)
        a.fill = PatternFill("solid", start_color="0A0A0A")
        a.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    else:
        a.font = Font(name="Arial", size=10, bold=True)
        b.font = Font(name="Arial", size=10)
    a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    legend.row_dimensions[i].height = 28 if body else 22

# Save
wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
print(f"Rows: {len(ROWS)}")
